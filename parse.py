from openpyxl import load_workbook
from openpyxl import Workbook
import helper
import constants

class Responses:

    def __init__(self, fname, code, jury):
        self.fname = fname
        self.CODE = code
        self.JURY = jury
        self.juryToStudents = {}
        self.parsed = []
        self.emailToApp = {}
        self.juryToGrToData = {}
        self.finalQueue = {}

        # screw it, hardcoding is king
        self.PROBCOLS = 'GHIJKLMNOP'

    def _parse_results(self):
        wb = load_workbook(self.fname)
        ws = wb['sheet']
        row = 2
        while True:
            if ws['A'+str(row)].value is None:
                break
            student = {}
            for col, param in self.CODE.items():
                if (response:=ws[col+str(row)].value) is not None and response != '-' and response != '.' and response != 'Нет вопросов':
                    student[param] = response
            self.parsed.append(student)
            row += 1

    def _find_uniques(self):
        for student in self.parsed:
            if (email:=student['email']) in self.emailToApp:
                print('someone has applied more than once')
            if email not in self.emailToApp:
                self.emailToApp[email] = {}
            for param, data in student.items():
                if param != 'email':
                    if param not in self.emailToApp[email]:
                        self.emailToApp[email][param] = []
                    self.emailToApp[email][param].append(data)

    def _make_queue_for_jury(self):
        for email, data in self.emailToApp.items():
            grade = data['grade'][0]
            for probCol in self.PROBCOLS:
                if (prob:=self.CODE[probCol]) in data:
                    studResponse = data[prob]
                    # print(grade, email, probCol)
                    jury = self.JURY[grade][prob]
                    # print(jury)
                    if jury not in self.juryToStudents:
                        self.juryToStudents[jury] = {}
                    if email not in self.juryToStudents[jury]:
                        self.juryToStudents[jury][email] = {}
                    for param, values in data.items():
                        if param not in self.juryToStudents[jury][email]:
                            self.juryToStudents[jury][email][param] = []
                        self.juryToStudents[jury][email][param].append(values)
                    # self.juryToStudents[jury].append(data)

    def _summary_for_jury(self):
        juryToGrToNum = {}
        for jury, emailToParam in self.juryToStudents.items():
            print(jury, len(emailToParam))
            if jury not in juryToGrToNum:
                juryToGrToNum[jury] = {}
            for email, paramToData in emailToParam.items():
                grade = paramToData['grade'][0][0] #hardcoding but screw it
                if grade not in juryToGrToNum[jury]:
                    juryToGrToNum[jury][grade] = []
                juryToGrToNum[jury][grade].append(paramToData)
        
        for jury, grToData in juryToGrToNum.items():
            for gr, data in grToData.items():
                print(jury, gr, len(data))
        
        self.juryToGrToData = juryToGrToNum

    def _create_the_queue(self):
        wb = Workbook()
        for grade in (9, 10, 11):
            wb.create_sheet(f"{grade} класс")
            ws = wb[f"{grade} класс"]
            col = 'A'
            for jury, grToData in self.juryToGrToData.items():
                row = 1
                ws[col + str(row)] = jury
                row += 1
                if grade in grToData:
                    for student in grToData[grade]:
                        # print(student['name'])
                        ws[col + str(row)] = row - 1 
                        ws[helper.getNextCol(col) + str(row)] = student['name'][0][0] 
                        row += 1 
                col = helper.getNextCol(col)
                col = helper.getNextCol(col)
        wb.save('queue.xlsx')

    def _is_there_a_conflict(self, data):
        for grade, juryToStudents in data.items():
            for jury, students in juryToStudents.items():
                for i, student in enumerate(students):
                    for jury2, students2 in juryToStudents.items():
                        if jury2 != jury:
                            if len(students2) > 0 and i < len(students2):
                                if students2[i] == student:
                                    return True
        return False

    def _is_there_a_conflict_EXTRA(self, data):
        for grade, juryToStudents in data.items():
            for jury, students in juryToStudents.items():
                for i, student in enumerate(students):
                    for jury2, students2 in juryToStudents.items():
                        if jury2 != jury:
                            if len(students2) > 0 and i < len(students2):
                                if students2[i] == student:
                                    return True
                            if len(students2) > 0 and i < len(students2) - 2:
                                if students2[i+1] == student:
                                    return True
        return False

    def _remove_repetitions_in_queue(self):
        wb = load_workbook('queue.xlsx')
        grToData = {}
        for grade in (9, 10, 11):
            ws = wb[f"{grade} класс"]
            juryToStuds = {}
            col = 'A'
            while True:
                if (jury := ws[col+str(1)].value) is None:
                    break

                row = 1
                
                if jury not in juryToStuds:
                    juryToStuds[jury] = []
                    row = 2
                    while True:
                        if (name:=ws[helper.getNextCol(col)+str(row)].value) is None:
                            break
                        juryToStuds[jury].append(name)
                        row += 1
                grToData[grade] = juryToStuds
                col = helper.getNextCol(col)
                col = helper.getNextCol(col)
        
        cntr = 0
        while True:
            cntr += 1
            # print(cntr)
            # anotherIter = False
            conflict = self._is_there_a_conflict(grToData)
            # conflict = self._is_there_a_conflict_EXTRA(grToData) # failure
            if not conflict:
                break
            anotherIter = False
            # if not anotherIter:
            for grade, juryToStudents in grToData.items():
                # if not anotherIter:
                for jury, students in juryToStudents.items():
                    # if not anotherIter:
                    for i, student in enumerate(students):
                        # if not anotherIter:
                        for jury2, students2 in juryToStudents.items():
                            if jury2 != jury:
                                if len(students2) > 0 and i < len(students2):
                                    if students2[i] == student:
                                        students = students[:i] + students[i+1:] + students[i:i+1]
                                        grToData[grade][jury] = students
                                        # anotherIter = True
                                        break
                                    # if not anotherIter:
                                    #     if len(students2) > 0 and i < len(students2) - 2:
                                    #         if students2[i+1] == student:
                                    #             students = students[:i+1] + students[i+2:] + students[i+1:i+2]
                                    #             grToData[grade][jury] = students
                                    #             anotherIter = True
                                    #             break
            # print('here')
        new_wb = Workbook()
        for grade, juryToStudents in grToData.items():
            new_wb.create_sheet(f"{grade} класс")
            ws = new_wb[f"{grade} класс"]
            col = 'A'
            for jury, students in juryToStudents.items():
                # print(students)
                row = 1
                ws[col + str(row)] = jury
                row += 1
                for student in students:
                    # print(student['name'])
                    ws[col + str(row)] = row - 1 
                    ws[helper.getNextCol(col) + str(row)] = student
                    row += 1 
                col = helper.getNextCol(col)
                col = helper.getNextCol(col)
        new_wb.save('queue_with_no_conflicts.xlsx')
        self.finalQueue = grToData

    def _print_jury_to_comments(self):
        juries = set()
        for grade, juryToStuds in self.finalQueue.items():
            juries = juries | set(juryToStuds.keys())
        for jury in juries:
            for grade, juryToStuds in self.finalQueue.items():
                f = open(f'{jury}-{grade}.txt', 'w')
                # wb.create_sheet(f"{grade} класс")
                # ws = wb[f"{grade} класс"]
                # row = 1
                for student in juryToStuds[jury]:
                    f.write('----------------------\n')
                    f.write(student+'\n')
                    for data in self.juryToGrToData[jury][grade]:
                        if data['name'][0][0] == student:
                            for param, response in data.items():
                                if param in constants.P_TO_JURY[grade]:
                                    if constants.P_TO_JURY[grade][param] == jury:
                                        f.write(f"--{param}\n")
                                        f.write(str(response) + '\n\n')



            # wb.save(f"{jury}.xlsx")

    def main(self):
        self._parse_results()
        self._find_uniques() #oh god why
        self._make_queue_for_jury()
        self._summary_for_jury()
        self._create_the_queue()
        self._remove_repetitions_in_queue()
        self._print_jury_to_comments()

# assume sheet is names "sheet"
chObj = Responses('chemistry.xlsx', constants.COL_TO_PARAM, constants.P_TO_JURY)
chObj.main()